import json
import pandas as pd
import argparse
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import *
import time

from tot.tasks import get_task

excel_file = 'analysis.xlsx'
task_name = ['k12b8', 'k8b5', 'k5b5', 'k5b3']
table_name = ['theoretical' , 'actual', 'traversal nodes', 'no ans in tree', 'error cot', 'wrong path']
algorithm_name = ['bfs', 'dfs+sd']
all_pos_table = {}

def parse_args():
    args = argparse.ArgumentParser()
    args.add_argument('--task', type=str, required=True, choices=['game24', 'text', 'crosswords'])
    args.add_argument('--task_start_index', type=int, default=900)
    args.add_argument('--task_end_index', type=int, default=1000)
    args.add_argument('--algorithm', type=str, required=True, choices=['bfs', 'dfs+sd', 'dfs+ksd', 'whole_tree']) # (bfs, dfs+sd, dfs+ksd, whole_tree)
    args.add_argument('--n_select_sample', type=int, default=1) # b
    args.add_argument('--k', type = int, default = 1) # k
    args.add_argument('--name_of_task', type=str, default='default')
    args = args.parse_args()
    return args

def Init(excel_path):
    print(os.listdir(excel_path))
    if os.path.exists(os.path.join(excel_path, excel_file)) == False:
        print('create a new excel file')
        wb = Workbook()
        wb.save(excel_file)
    else:
        print('already have excel file')
        wb = load_workbook(os.path.join(excel_path, excel_file))
    return wb

def error_cot(states, task):
    theoretical = 0
    actual = 0
    correct_list = [0 for _ in range(args.task_end_index - args.task_start_index)]
    for state in states:
        ans = state['ys'][0]
        if 'Answer' not in ans:
            continue
        # if has answer means final number == 24
        theoretical += 1
        correct_list[state['idx'] - 900] = 1
        correct = task.test_output(state['idx'], ans)['r']
        if correct == 1:
            actual += 1
    return theoretical, actual, correct_list

def no_ans_in_base(tree):
    has_ans_list = [0 for _ in range(args.task_end_index - args.task_start_index)]
    for state in tree:
        for step in state['steps']:
            if step['step'] == 2:
                leaf_nodes = step['ys']
        # print(leaf_nodes)
        for node in leaf_nodes:
            # print(node)
            ans = node[1].split('\n')[-2]
            # print(ans)
            if 'left: ' not in ans:
                    continue
            ans = ans.split('left: ')[-1].replace(')', '').strip()
            # print(ans)
            if ans == '24':
                has_ans_list[state['idx'] - 900] = 1

    no_ans_count = sum(1 for num in has_ans_list if num == 0)
    return has_ans_list, no_ans_count

def wrong_path(has_ans_list, correct_list, impossible):
    wrong_path_count = 0
    for i in range(args.task_end_index - args.task_start_index):
        if has_ans_list[i] == 0 and correct_list[i] == 1:
            impossible += 1
        elif has_ans_list[i] == 1 and correct_list[i] == 0:
            print(i + 900)
            wrong_path_count += 1
    return wrong_path_count, impossible

def format_border(ws, start_row, end_row, start_col, end_col):
    # left
    for row in range(start_row, end_row + 1):
        ws.cell(row, start_col).border = Border(left = Side(style = 'medium'))
    # right
    for row in range(start_row, end_row + 1):
        ws.cell(row, end_col + 1).border = Border(left = Side(style = 'medium'))
    # top
    for col in range(start_col, end_col + 1):
        ws.cell(start_row, col).border = Border(top = Side(style = 'medium'))
    # bottom
    for col in range(start_col, end_col + 1):
        ws.cell(end_row + 1, col).border = Border(top = Side(style = 'medium'))

def multi_table(ws, pos, pos_table):
    col = 1
    for name in table_name:
        pos_table[name] = {}
        start_col = col
        end_pos, col = create_table(ws, name, pos, col, pos_table[name])
        format_border(ws, pos, end_pos, start_col, col)
        col += 2
    return end_pos, col

def create_table(ws, table_name, pos, col, pos_table):
    start_col = col
    ws.cell(pos, col).value = table_name
    col += 1
    for i in range(3):
        ws.cell(pos, col).value = i
        col += 1
    ws.cell(pos, col).value = 'avg'
    for name in task_name:
        pos += 1
        ws.cell(pos, start_col).value = name
        pos_table[name] = (pos, start_col)
    return pos, col

def format_table(ws):
    pos = 1
    col = 1
    for name in algorithm_name:
        all_pos_table[name] = {}
        ws.cell(pos, col).value = name
        pos, _ = multi_table(ws, pos + 1, all_pos_table[name])
        pos += 1

def append_data(ws, algorithm_name, task_name, theoretical, actual, traversal_nodes, no_ans_in_base, wrong_path, error_cot):
    # theoretical
    pos, col = all_pos_table[algorithm_name]['theoretical'][task_name]
    for i in range(3):
        if ws.cell(pos, col + i).value == None:
            ws.cell(pos, col + i).value = theoretical
            break
    cal_avg(ws, pos, col)
    # actual
    pos, col = all_pos_table[algorithm_name]['actual'][task_name]
    for i in range(3):
        if ws.cell(pos, col + i).value == None:
            ws.cell(pos, col + i).value = actual
            break
    cal_avg(ws, pos, col)
    # traversal_nodes
    pos, col = all_pos_table[algorithm_name]['traversal nodes'][task_name]
    for i in range(3):
        if ws.cell(pos, col + i).value == None:
            ws.cell(pos, col + i).value = traversal_nodes
            break
    cal_avg(ws, pos, col)
    # no ans in tree
    pos, col = all_pos_table[algorithm_name]['no ans in tree'][task_name]
    for i in range(3):
        if ws.cell(pos, col + i).value == None:
            ws.cell(pos, col + i).value = no_ans_in_base
            break
    cal_avg(ws, pos, col)
    # wrong path
    pos, col = all_pos_table[algorithm_name]['wrong path'][task_name]
    for i in range(3):
        if ws.cell(pos, col + i).value == None:
            ws.cell(pos, col + i).value = wrong_path
            break
    cal_avg(ws, pos, col)
    # error cot
    pos, col = all_pos_table[algorithm_name]['error cot'][task_name]
    for i in range(3):
        if ws.cell(pos, col + i).value == None:
            ws.cell(pos, col + i).value = error_cot
            break
    cal_avg(ws, pos, col)

def cal_avg(ws, pos, col):
    sum = 0
    data_num = 3
    for i in range(1, 4):
        if ws.cell(pos, col + i).value == None:
            data_num -= 1
            continue
        sum += ws.cell(pos, col + i).value
    ws.cell(pos, col + 4).value = sum / data_num

def run(args):
    # load excel file
    excel_path = f'./logs/{args.task}/{args.name_of_task}'
    wb = Init(excel_path)
    ws = wb.active

    task = get_task(args.task)
    path = f'./logs/{args.task}/{args.name_of_task}/k{args.k}b{args.n_select_sample}'
    dir_list = os.listdir(path)
    dir_list = [item for item in dir_list if os.path.isdir(os.path.join(path, item))]
    for name in dir_list:
        folder_name = f'./logs/{args.task}/{args.name_of_task}/k{args.k}b{args.n_select_sample}/{name}'
        print(folder_name)
        file_name = os.path.join(folder_name, f'{args.name_of_task}_start{args.task_start_index}_end{args.task_end_index}_{args.algorithm}_0.json')
        # print(file_name)
        with open(file_name, 'r') as file:
             data = json.load(file)

        if args.algorithm == 'whole_tree':
            base = [data[i] for i in range(len(data)) if i % 3 == 0]
            print('\n'.join(map(str, base)))

            bfs = [data[i] for i in range(len(data)) if i % 3 == 1]
            # print('\n'.join(map(str, bfs)))

            dfs = [data[i] for i in range(len(data)) if i % 3 == 2]
            # print('\n'.join(map(str, dfs)))

            bfs_traversal_nodes = sum([state['traversal_nodes'] for state in bfs]) / len(bfs)
            dfs_traversal_nodes = sum([state['traversal_nodes'] for state in dfs]) / len(dfs)

            # error cot
            # bfs
            bfs_theoretical, bfs_actual, bfs_correct_list = error_cot(bfs, task)
            print(f'bfs (theoretical, actual): ({bfs_theoretical, bfs_actual})')
            # dfs
            dfs_theoretical, dfs_actual, dfs_correct_list = error_cot(dfs, task)
            print(f'dfs (theoretical, actual): ({dfs_theoretical, dfs_actual})')

            # no ans in base
            has_ans_list, no_ans_count = no_ans_in_base(base)
            print(has_ans_list)
            count = 0
            for i in has_ans_list:
                if i == 0:
                    print(count)
                count += 1
            print('no ans tree count: ' + str(no_ans_count))

            impossible = 0
            # wrong path bfs
            bfs_wrong_path_count, impossible = wrong_path(has_ans_list, bfs_correct_list, impossible)
            print('bfs wrong path count: ' + str(bfs_wrong_path_count))
            
            # wrong path dfs
            dfs_wrong_path_count, impossible = wrong_path(has_ans_list, dfs_correct_list, impossible)
            print('dfs wrong path count: ' + str(dfs_wrong_path_count))
            print('impossible: ' + str(impossible))
 
            # output as excel file
            format_table(ws)
            append_data(ws, 'bfs', f'k{args.k}b{args.n_select_sample}', bfs_theoretical, bfs_actual, bfs_traversal_nodes, no_ans_count, bfs_wrong_path_count, bfs_theoretical - bfs_actual)
            append_data(ws, 'dfs+sd', f'k{args.k}b{args.n_select_sample}', dfs_theoretical, dfs_actual, dfs_traversal_nodes, no_ans_count, dfs_wrong_path_count, dfs_theoretical - dfs_actual)

    wb.save(os.path.join(excel_path, excel_file))
    print('output as excel file')


if __name__ == '__main__':
    args = parse_args()
    print(args)
    run(args)